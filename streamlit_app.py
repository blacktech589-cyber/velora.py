"""Binomo için CSV tabanlı araştırma ve Excel sinyal uygulaması.

Bu yazılım yatırım tavsiyesi veya otomatik işlem aracı değildir. Binomo'dan
dışa aktarılan mum verisini analiz eder; platform hesabına bağlanmaz.
"""
from __future__ import annotations

import io
import hashlib
import json
import logging
import os
import time
import warnings
from datetime import datetime
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import Request, urlopen

# Streamlit Cloud'da BLAS/joblib'in aşırı işçi oluşturmasını engeller.
os.environ.setdefault("OMP_NUM_THREADS", "1")
os.environ.setdefault("OPENBLAS_NUM_THREADS", "1")
os.environ.setdefault("MKL_NUM_THREADS", "1")
os.environ.setdefault("NUMEXPR_NUM_THREADS", "1")
warnings.filterwarnings(
    "ignore",
    message=r".*sklearn\.utils\.parallel\.delayed.*",
    category=UserWarning,
)

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from sklearn.ensemble import HistGradientBoostingClassifier
from sklearn.impute import SimpleImputer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, precision_score, recall_score
from sklearn.neural_network import MLPClassifier
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler

# Community Cloud güvenli sürümünde ağır TensorFlow çalışma zamanı yüklenmez.
# DeepMLP, çok katmanlı sinir ağı modeli olarak ensemble içinde çalışır.
TF_AVAILABLE = False


APP_DIR = Path(__file__).resolve().parent
MODEL_DIR = APP_DIR / "models"
OUTPUT_FILE = APP_DIR / "binomo_sinyalleri.xlsx"
AUTO_CSV_FILE = APP_DIR / "hazir_veri.csv"
LOG_FILE = APP_DIR / "hata_log.txt"
MODEL_DIR.mkdir(exist_ok=True)
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    encoding="utf-8",
)

FEATURES = [
    # Getiriler ve trend
    "return_1", "return_3", "return_5", "return_10",
    "sma_ratio_5", "sma_ratio_10", "sma_ratio_20", "sma_ratio_50",
    "ema_ratio_5", "ema_ratio_9", "ema_ratio_12", "ema_ratio_15",
    "ema_ratio_21", "ema_ratio_26", "ema_ratio_50",
    "ema_15_slope", "ema_15_26_spread",
    # Momentum
    "rsi_7", "rsi_14", "rsi_21", "stoch_k", "stoch_d",
    "williams_r", "roc_10", "cci_20",
    # Trend gücü
    "macd", "macd_signal", "macd_histogram", "adx_14", "plus_di", "minus_di",
    # Oynaklık ve fiyat konumu
    "volatility_10", "volatility_20", "atr_14", "natr_14",
    "bb_position", "bb_width", "bb_percent_b", "range_ratio",
    # Hacim
    "volume_change", "obv_change", "mfi_14", "vwap_ratio",
]

MARKET_SYMBOLS = {
    "Crypto IDX (piyasa vekili: Bitcoin)": "BTC-USD",
    "Bitcoin (OTC vekili)": "BTC-USD",
    "Ethereum (OTC vekili)": "ETH-USD",
    "Solana (OTC vekili)": "SOL-USD",
    "FC Barcelona Token (OTC vekili)": "BAR-USD",
    "Cardano (OTC vekili)": "ADA-USD",
    "Chainlink (OTC vekili)": "LINK-USD",
    "Bitcoin Cash (OTC vekili)": "BCH-USD",
    "Kusama (OTC vekili)": "KSM-USD",
    "Aave (OTC vekili)": "AAVE-USD",
    "PancakeSwap (OTC vekili)": "CAKE-USD",
    "Uniswap (OTC vekili)": "UNI-USD",
    "EUR/USD": "EURUSD=X",
    "EUR/USD (OTC vekili)": "EURUSD=X",
    "GBP/USD": "GBPUSD=X",
    "GBP/USD (OTC vekili)": "GBPUSD=X",
    "USD/JPY": "JPY=X",
    "USD/JPY (OTC vekili)": "JPY=X",
    "USD/CHF": "CHF=X",
    "AUD/USD": "AUDUSD=X",
    "AUD/USD (OTC vekili)": "AUDUSD=X",
    "USD/CAD": "CAD=X",
    "USD/CAD (OTC vekili)": "CAD=X",
    "NZD/USD": "NZDUSD=X",
    "NZD/USD (OTC vekili)": "NZDUSD=X",
    "EUR/GBP": "EURGBP=X",
    "EUR/GBP (OTC vekili)": "EURGBP=X",
    "EUR/JPY": "EURJPY=X",
    "GBP/JPY": "GBPJPY=X",
    "GBP/JPY (OTC vekili)": "GBPJPY=X",
    "GBP/CHF (OTC vekili)": "GBPCHF=X",
    "CHF/JPY (OTC vekili)": "CHFJPY=X",
    "EUR/CAD (OTC vekili)": "EURCAD=X",
    "AUD/CAD (OTC vekili)": "AUDCAD=X",
    "GBP/NZD (OTC vekili)": "GBPNZD=X",
    "AUD/JPY": "AUDJPY=X",
    "EUR/CHF": "EURCHF=X",
    "GBP/AUD": "GBPAUD=X",
    "Bitcoin/USD": "BTC-USD",
    "Ethereum/USD": "ETH-USD",
    "Solana/USD": "SOL-USD",
    "Gold": "GC=F",
    "Silver": "SI=F",
    "Oil (WTI)": "CL=F",
    "Natural Gas": "NG=F",
    "S&P 500": "^GSPC",
    "NASDAQ 100": "^NDX",
    "DAX 40": "^GDAXI",
    "Apple": "AAPL",
    "Microsoft": "MSFT",
    "Nvidia": "NVDA",
    "Tesla": "TSLA",
    "Amazon": "AMZN",
    "Meta": "META",
    "Google": "GOOGL",
}

# Binomo ekranında görülebilen normal Forex varlıkları. Harici veri
# servisindeki karşılıkları örneğin CAD/JPY -> CADJPY=X biçimindedir.
FOREX_PAIRS = [
    "EUR/USD", "CAD/JPY", "AUD/ZAR", "GBP/NOK", "EUR/HUF", "GBP/SGD",
    "CAD/DKK", "EUR/ILS", "USD/MXN", "AUD/USD", "EUR/NOK", "AUD/JPY",
    "CAD/MXN", "EUR/ZAR", "GBP/HKD", "CHF/JPY", "USD/SEK", "CAD/SEK",
    "USD/ILS", "CAD/NOK", "EUR/SEK", "AUD/SGD", "USD/PLN", "CAD/SGD",
    "GBP/MXN", "CHF/DKK", "AUD/DKK", "EUR/SGD", "GBP/CZK", "GBP/PLN",
    "AUD/HUF", "EUR/NZD", "EUR/MXN", "NZD/JPY", "USD/JPY", "CHF/NOK",
    "USD/HUF", "AUD/NOK", "GBP/DKK", "GBP/TRY", "AUD/SEK", "CHF/PLN",
    "USD/CAD", "CHF/SEK", "NOK/JPY", "NOK/SEK", "CHF/SGD", "NZD/CAD",
    "NZD/DKK", "NZD/NOK", "GBP/NZD", "NZD/HUF", "NZD/SEK", "NZD/SGD",
    "SEK/JPY", "SGD/HKD", "SGD/JPY", "USD/CHF", "AUD/CAD", "USD/CZK",
    "NZD/CHF", "GBP/CAD", "AUD/NZD", "USD/ZAR", "AUD/CHF", "GBP/USD",
    "EUR/PLN", "GBP/HUF", "ZAR/JPY", "CHF/HUF", "GBP/CHF", "USD/DKK",
    "EUR/CHF", "GBP/AUD", "EUR/GBP", "EUR/JPY",
]
for forex_pair in FOREX_PAIRS:
    MARKET_SYMBOLS.setdefault(forex_pair, forex_pair.replace("/", "") + "=X")

# Binomo ekranındaki adlarla otomatik taranacak öncelikli varlıklar.
# OTC fiyatları halka açık olmadığı için karşılarında normal piyasa vekilleri var.
BINOMO_ASSET_SYMBOLS = {
    "Crypto IDX": "BTC-USD",
    "Bitcoin (OTC)": "BTC-USD",
    "Ethereum (OTC)": "ETH-USD",
    "Solana (OTC)": "SOL-USD",
    "FC Barcelona Token (OTC)": "BAR-USD",
    "AUD/USD (OTC)": "AUDUSD=X",
    "Cardano (OTC)": "ADA-USD",
    "NZD/USD (OTC)": "NZDUSD=X",
    "GBP/CHF (OTC)": "GBPCHF=X",
    "Chainlink (OTC)": "LINK-USD",
    "CHF/JPY (OTC)": "CHFJPY=X",
    "EUR/CAD (OTC)": "EURCAD=X",
    "GBP/JPY (OTC)": "GBPJPY=X",
    "GBP/USD (OTC)": "GBPUSD=X",
    "EUR/GBP (OTC)": "EURGBP=X",
    "EUR/USD (OTC)": "EURUSD=X",
    "AUD/CAD (OTC)": "AUDCAD=X",
    "USD/JPY (OTC)": "JPY=X",
    "USD/CAD (OTC)": "CAD=X",
    "Bitcoin Cash (OTC)": "BCH-USD",
    "Kusama (OTC)": "KSM-USD",
    "Aave (OTC)": "AAVE-USD",
    "Pancake Swap (OTC)": "CAKE-USD",
    "Uniswap (OTC)": "UNI-USD",
    "GBP/NZD (OTC)": "GBPNZD=X",
    "AUD/JPY": "AUDJPY=X",
    "Gold": "GC=F",
}
MARKET_SYMBOLS.update(BINOMO_ASSET_SYMBOLS)
AUTO_SCAN_ASSETS = list(BINOMO_ASSET_SYMBOLS)

INTERVAL_PERIODS = {
    "1m": "7d",
    "5m": "60d",
    "15m": "60d",
    "30m": "60d",
    "1h": "2y",
    "1d": "10y",
}


@st.cache_data(ttl=20, show_spinner=False)
def download_market_data(symbol: str, interval: str) -> pd.DataFrame:
    """Ek paket gerektirmeden harici piyasa kaynağından OHLCV indirir."""
    encoded_symbol = urlencode({"symbol": symbol}).split("=", 1)[1]
    query = urlencode({
        "interval": interval,
        "range": INTERVAL_PERIODS[interval],
        "includePrePost": "false",
        "events": "div,splits",
    })
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{encoded_symbol}?{query}"
    request = Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": "application/json",
        },
    )
    try:
        with urlopen(request, timeout=30) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except Exception as exc:
        raise ValueError(f"Piyasa veri servisine erişilemedi: {exc}") from exc

    chart = payload.get("chart", {})
    if chart.get("error"):
        raise ValueError(str(chart["error"]))
    results = chart.get("result") or []
    if not results:
        raise ValueError(f"{symbol} için çevrim içi veri bulunamadı.")
    result = results[0]
    timestamps = result.get("timestamp") or []
    quotes = ((result.get("indicators") or {}).get("quote") or [{}])[0]
    if not timestamps or not quotes:
        raise ValueError(f"{symbol} için mum verisi boş döndü.")
    size = len(timestamps)
    downloaded = pd.DataFrame({
        "time": pd.to_datetime(timestamps, unit="s", utc=True),
        "open": quotes.get("open", [None] * size),
        "high": quotes.get("high", [None] * size),
        "low": quotes.get("low", [None] * size),
        "close": quotes.get("close", [None] * size),
        "volume": quotes.get("volume", [0] * size),
    })
    return normalize_columns(downloaded)


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    aliases = {
        "timestamp": "time", "datetime": "time", "date": "time",
        "tarih": "time", "zaman": "time", "açılış": "open", "acilis": "open",
        "yüksek": "high", "yuksek": "high", "düşük": "low", "dusuk": "low",
        "kapanış": "close", "kapanis": "close", "hacim": "volume",
    }
    out = df.copy()
    out.columns = [aliases.get(str(c).strip().lower(), str(c).strip().lower()) for c in out.columns]
    required = {"time", "open", "high", "low", "close"}
    missing = required.difference(out.columns)
    if missing:
        raise ValueError("Eksik sütun(lar): " + ", ".join(sorted(missing)))
    if "volume" not in out:
        out["volume"] = 0.0
    out["time"] = pd.to_datetime(out["time"], errors="coerce", utc=True)
    for col in ["open", "high", "low", "close", "volume"]:
        out[col] = pd.to_numeric(out[col], errors="coerce")
    out = out.dropna(subset=["time", "open", "high", "low", "close"])
    out = out.sort_values("time").drop_duplicates("time").reset_index(drop=True)
    if len(out) < 150:
        raise ValueError("Eğitim için en az 150 geçerli mum gerekir.")
    return out


def rsi(close: pd.Series, period: int = 14) -> pd.Series:
    delta = close.diff()
    gain = delta.clip(lower=0).ewm(alpha=1 / period, adjust=False).mean()
    loss = (-delta.clip(upper=0)).ewm(alpha=1 / period, adjust=False).mean()
    rs = gain / loss.replace(0, np.nan)
    return (100 - 100 / (1 + rs)).fillna(50) / 100


def true_range(frame: pd.DataFrame) -> pd.Series:
    previous_close = frame["close"].shift(1)
    return pd.concat([
        frame["high"] - frame["low"],
        (frame["high"] - previous_close).abs(),
        (frame["low"] - previous_close).abs(),
    ], axis=1).max(axis=1)


def directional_index(frame: pd.DataFrame, period: int = 14):
    up = frame["high"].diff()
    down = -frame["low"].diff()
    plus_dm = pd.Series(np.where((up > down) & (up > 0), up, 0.0), index=frame.index)
    minus_dm = pd.Series(np.where((down > up) & (down > 0), down, 0.0), index=frame.index)
    atr = true_range(frame).ewm(alpha=1 / period, adjust=False).mean()
    plus_di = 100 * plus_dm.ewm(alpha=1 / period, adjust=False).mean() / atr.replace(0, np.nan)
    minus_di = 100 * minus_dm.ewm(alpha=1 / period, adjust=False).mean() / atr.replace(0, np.nan)
    dx = 100 * (plus_di - minus_di).abs() / (plus_di + minus_di).replace(0, np.nan)
    adx = dx.ewm(alpha=1 / period, adjust=False).mean()
    return adx / 100, plus_di / 100, minus_di / 100


def money_flow_index(frame: pd.DataFrame, period: int = 14) -> pd.Series:
    typical = (frame["high"] + frame["low"] + frame["close"]) / 3
    flow = typical * frame["volume"]
    positive = flow.where(typical.diff() > 0, 0.0).rolling(period).sum()
    negative = flow.where(typical.diff() < 0, 0.0).rolling(period).sum()
    ratio = positive / negative.replace(0, np.nan)
    return (100 - 100 / (1 + ratio)).fillna(50) / 100


def add_features(df: pd.DataFrame, horizon: int) -> pd.DataFrame:
    x = df.copy()
    close = x["close"]
    for n in [1, 3, 5, 10]:
        x[f"return_{n}"] = close.pct_change(n)
    for n in [5, 10, 20, 50]:
        x[f"sma_ratio_{n}"] = close / close.rolling(n).mean() - 1
    emas = {}
    for n in [5, 9, 12, 15, 21, 26, 50]:
        emas[n] = close.ewm(span=n, adjust=False).mean()
        x[f"ema_ratio_{n}"] = close / close.ewm(span=n, adjust=False).mean() - 1
    x["ema_15_slope"] = emas[15].pct_change(3)
    x["ema_15_26_spread"] = emas[15] / emas[26] - 1
    returns = close.pct_change()
    x["volatility_10"] = returns.rolling(10).std()
    x["volatility_20"] = returns.rolling(20).std()
    x["rsi_7"] = rsi(close, 7)
    x["rsi_14"] = rsi(close)
    x["rsi_21"] = rsi(close, 21)
    ema12, ema26 = emas[12], emas[26]
    x["macd"] = (ema12 - ema26) / close
    x["macd_signal"] = (ema12 - ema26).ewm(span=9, adjust=False).mean() / close
    x["macd_histogram"] = x["macd"] - x["macd_signal"]
    mid = close.rolling(20).mean()
    deviation = close.rolling(20).std()
    upper, lower = mid + 2 * deviation, mid - 2 * deviation
    band_range = (upper - lower).replace(0, np.nan)
    x["bb_position"] = (close - lower) / band_range
    x["bb_percent_b"] = x["bb_position"]
    x["bb_width"] = band_range / mid.replace(0, np.nan)
    low14 = x["low"].rolling(14).min()
    high14 = x["high"].rolling(14).max()
    x["stoch_k"] = (close - low14) / (high14 - low14).replace(0, np.nan)
    x["stoch_d"] = x["stoch_k"].rolling(3).mean()
    x["williams_r"] = (close - high14) / (high14 - low14).replace(0, np.nan)
    x["roc_10"] = close.pct_change(10)
    typical = (x["high"] + x["low"] + close) / 3
    mean_deviation = typical.rolling(20).apply(
        lambda values: np.mean(np.abs(values - values.mean())), raw=True
    )
    x["cci_20"] = (typical - typical.rolling(20).mean()) / (0.015 * mean_deviation).replace(0, np.nan)
    atr = true_range(x).ewm(alpha=1 / 14, adjust=False).mean()
    x["atr_14"] = atr / close.replace(0, np.nan)
    x["natr_14"] = x["atr_14"]
    x["adx_14"], x["plus_di"], x["minus_di"] = directional_index(x)
    x["range_ratio"] = (x["high"] - x["low"]) / close.replace(0, np.nan)
    # Forex kaynaklarında gerçek hacim bulunmayabilir. Böyle bir durumda
    # hacim-tabanlı özellikleri sıfır/nötr üretmek eğitim satırlarını korur.
    has_volume = bool((x["volume"].fillna(0) > 0).any())
    effective_volume = x["volume"].fillna(0) if has_volume else pd.Series(1.0, index=x.index)
    x["volume_change"] = effective_volume.pct_change().replace([np.inf, -np.inf], np.nan).fillna(0)
    direction = np.sign(close.diff()).fillna(0)
    obv = (direction * effective_volume).cumsum()
    x["obv_change"] = obv.diff(5) / (obv.abs().rolling(20).mean() + 1e-9)
    if has_volume:
        x["mfi_14"] = money_flow_index(x)
    else:
        x["mfi_14"] = 0.5
    cumulative_volume = effective_volume.cumsum()
    vwap = (typical * effective_volume).cumsum() / cumulative_volume
    x["vwap_ratio"] = close / vwap.replace(0, np.nan) - 1
    future_return = close.shift(-horizon) / close - 1
    x["target"] = np.where(future_return.notna(), (future_return > 0).astype(int), np.nan)
    return x.replace([np.inf, -np.inf], np.nan)


def sequences(frame: pd.DataFrame, lookback: int, max_sequences: int = 6000):
    values = frame[FEATURES].to_numpy(dtype=np.float32)
    labels = frame["target"].to_numpy()
    xs, ys, rows = [], [], []
    eligible = np.array([
        end for end in range(lookback - 1, len(frame))
        if not np.isnan(labels[end])
    ], dtype=int)
    if len(eligible) > max_sequences:
        positions = np.linspace(
            0, len(eligible) - 1, num=max_sequences, dtype=int
        )
        eligible = eligible[positions]
    for end in eligible:
        window = values[end - lookback + 1:end + 1]
        med = np.nanmedian(window, axis=0)
        # Tamamen boş kalan opsiyonel bir gösterge nötr (0) kabul edilir.
        med = np.nan_to_num(med, nan=0.0, posinf=0.0, neginf=0.0)
        window = np.where(np.isnan(window), med, window)
        window = np.nan_to_num(window, nan=0.0, posinf=0.0, neginf=0.0)
        xs.append(window)
        ys.append(int(labels[end]))
        rows.append(end)
    return np.asarray(xs), np.asarray(ys), rows


def train_and_predict(frame, lookback, epochs, max_windows=6000):
    X, y, row_ids = sequences(
        frame, lookback, max_sequences=min(6000, int(max_windows))
    )
    if len(X) < 100:
        raise ValueError(
            f"Yalnızca {len(X)} eğitim penceresi oluştu. Daha uzun bir mum "
            "aralığı seçin (15m, 30m veya 1h önerilir)."
        )
    if len(np.unique(y)) < 2:
        raise ValueError(
            "Seçilen dönemde fiyat yalnızca tek yönde hareket etmiş. "
            "Mum aralığını veya tahmin ufkunu değiştirin."
        )
    # Cloud belleğini korurken en yeni piyasa rejimine öncelik verir.
    split = int(len(X) * .8)
    if split < 50 or len(X) - split < 20:
        raise ValueError("Zaman bazlı test bölümü için daha fazla mum gerekiyor.")
    X_train, X_test, y_train, y_test = X[:split], X[split:], y[:split], y[split:]
    flat_train = X_train.reshape(len(X_train), -1)
    flat_test = X_test.reshape(len(X_test), -1)
    latest_seq = X[-1:]
    probabilities, latest_probs, names = [], [], []

    classical = [
        ("DeepMLP", MLPClassifier(
            hidden_layer_sizes=(64, 32), activation="relu",
            alpha=0.001, batch_size=64,
            max_iter=max(40, int(epochs) * 3),
            early_stopping=True, validation_fraction=0.15,
            n_iter_no_change=10, random_state=42,
        )),
        ("HistGradientBoosting", HistGradientBoostingClassifier(
            max_iter=160, learning_rate=0.05, max_leaf_nodes=31,
            l2_regularization=0.1, early_stopping=True, random_state=42,
        )),
        ("LogisticRegression", LogisticRegression(
            C=0.5, max_iter=500, class_weight="balanced", random_state=42,
        )),
    ]
    for name, estimator in classical:
        pipe = Pipeline([("imputer", SimpleImputer()), ("scale", StandardScaler()), ("model", estimator)])
        pipe.fit(flat_train, y_train)
        probabilities.append(pipe.predict_proba(flat_test)[:, 1])
        latest_probs.append(
            float(pipe.predict_proba(latest_seq.reshape(1, -1))[0, 1])
        )
        names.append(name)

    # Her modelin yakın geçmiş test başarısına göre dinamik ağırlıklandırma.
    model_scores = [
        accuracy_score(y_test, (model_probability >= 0.5).astype(int))
        for model_probability in probabilities
    ]
    model_weights = np.array(
        [max(0.05, score - 0.45) for score in model_scores], dtype=float
    )
    model_weights /= model_weights.sum()
    ensemble = np.average(np.vstack(probabilities), axis=0, weights=model_weights)
    pred = (ensemble >= .5).astype(int)
    disagreement = float(np.mean(np.std(np.vstack(probabilities), axis=0)))
    metrics = {
        "Doğruluk": accuracy_score(y_test, pred),
        "Precision": precision_score(y_test, pred, zero_division=0),
        "Recall": recall_score(y_test, pred, zero_division=0),
        "Test örneği": len(y_test),
        "Model anlaşmazlığı": round(disagreement, 4),
    }

    probability = float(np.average(latest_probs, weights=model_weights))
    model_details = {
        name: {
            "test_doğruluğu": round(float(score), 4),
            "zeka_ağırlığı": round(float(weight), 4),
            "yukarı_olasılığı": round(float(latest), 4),
        }
        for name, score, weight, latest in zip(
            names, model_scores, model_weights, latest_probs
        )
    }
    return probability, metrics, names, model_details


def replace_legacy_wait_signals(frame: pd.DataFrame) -> pd.DataFrame:
    """Eski BEKLE kayıtlarını model olasılığına göre yön sinyaline çevirir."""
    result = frame.copy()
    if "Sinyal" not in result or "Model olasılığı" not in result:
        return result
    probability = pd.to_numeric(result["Model olasılığı"], errors="coerce")
    wait_mask = result["Sinyal"].astype(str).str.upper().eq("BEKLE")
    result.loc[wait_mask & probability.ge(0.5), "Sinyal"] = "YUKARI"
    result.loc[wait_mask & probability.lt(0.5), "Sinyal"] = "AŞAĞI"
    return result


def append_excel(record: dict, market_data: pd.DataFrame | None = None) -> bytes:
    new = pd.DataFrame([record])
    if OUTPUT_FILE.exists():
        old = pd.read_excel(OUTPUT_FILE, sheet_name="Sinyaller")
        data = pd.concat([old, new], ignore_index=True).tail(5000)
    else:
        data = new
    # Yeni sonuçla birlikte eski BEKLE satırlarını da Excel'de kalıcı düzelt.
    data = replace_legacy_wait_signals(data)
    # En yüksek güvenli işlemler her zaman ilk sırada görünür.
    data = data.drop(columns=["Öncelik"], errors="ignore")
    data["Güven"] = pd.to_numeric(data["Güven"], errors="coerce")
    payout_source = data.get(
        "Binomo ödeme oranı (%)",
        pd.Series(0.0, index=data.index),
    )
    payout = pd.to_numeric(payout_source, errors="coerce").fillna(0)
    data["Karşılaştırma puanı"] = (
        data["Güven"] * payout - (1 - data["Güven"]) * 100
    ).round(2)
    data = data.sort_values(
        ["Karşılaştırma puanı", "Güven", "Zaman"],
        ascending=[False, False, False],
        na_position="last",
    ).reset_index(drop=True)
    data.insert(0, "Öncelik", np.arange(1, len(data) + 1))
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        data.to_excel(writer, sheet_name="Sinyaller", index=False)
        ws = writer.sheets["Sinyaller"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            letter = col[0].column_letter
            ws.column_dimensions[letter].width = min(32, max(12, max(len(str(c.value or "")) for c in col) + 2))
        if ws.max_row > 1:
            ws.conditional_formatting.add(
                f"F2:F{ws.max_row}",
                ColorScaleRule(start_type="min", start_color="F8696B",
                               mid_type="percentile", mid_value=50, mid_color="FFEB84",
                               end_type="max", end_color="63BE7B"),
            )
            for cell in ws["E"][1:]:
                cell.number_format = "0.00%"
            for cell in ws["F"][1:]:
                cell.number_format = "0.00%"
        pd.DataFrame([{
            "Not": "Araştırma amaçlıdır; yatırım tavsiyesi veya kazanç garantisi değildir.",
            "Veri": "Kullanıcının yüklediği geçmiş OHLC mum verisi",
            "Doğrulama": "Son %20 veri, zaman sırası korunarak test edilir.",
        }]).to_excel(writer, sheet_name="Açıklama", index=False)
        if market_data is not None and not market_data.empty:
            export_data = market_data.tail(5000).copy()
            # Excel saat dilimli datetime değerlerini kabul etmez.
            export_data["time"] = export_data["time"].dt.tz_localize(None)
            export_data.to_excel(writer, sheet_name="Piyasa_Verisi", index=False)
    return OUTPUT_FILE.read_bytes()


def prioritized_signal_table() -> pd.DataFrame:
    """Excel geçmişini ekranda güven yüzdesine göre önceliklendirir."""
    if not OUTPUT_FILE.exists():
        return pd.DataFrame()
    history = pd.read_excel(OUTPUT_FILE, sheet_name="Sinyaller")
    history = replace_legacy_wait_signals(history)
    history["Güven"] = pd.to_numeric(history["Güven"], errors="coerce")
    payout_source = history.get(
        "Binomo ödeme oranı (%)",
        pd.Series(0.0, index=history.index),
    )
    payout = pd.to_numeric(payout_source, errors="coerce").fillna(0)
    # Bir birim risk için ikili işlem beklenen değeri:
    # kazanma_olasılığı * ödeme - kaybetme_olasılığı * 100
    history["Karşılaştırma puanı"] = (
        history["Güven"] * payout - (1 - history["Güven"]) * 100
    ).round(2)
    history = history.sort_values(
        ["Karşılaştırma puanı", "Güven", "Zaman"],
        ascending=[False, False, False],
        na_position="last",
    ).reset_index(drop=True)
    history["Öncelik"] = np.arange(1, len(history) + 1)
    history["Güven"] = history["Güven"] * 100
    if "Model olasılığı" in history:
        history["Model olasılığı"] = history["Model olasılığı"] * 100
    if "Test doğruluğu" in history:
        history["Test doğruluğu"] = history["Test doğruluğu"] * 100
    visible = [
        "Öncelik", "Varlık", "Sinyal", "Karşılaştırma puanı",
        "Güven yüzdesi", "Güven", "Binomo ödeme oranı (%)", "Model olasılığı",
        "Tahmin ufku (mum)", "Test doğruluğu", "Zaman",
    ]
    return history[[column for column in visible if column in history.columns]]


def render_clickable_ranking(table: pd.DataFrame, key: str):
    """Sıralamayı gösterir ve tıklanan işlemin ayrıntısını açar."""
    event = st.dataframe(
        table,
        use_container_width=True,
        hide_index=True,
        key=key,
        on_select="rerun",
        selection_mode="single-row",
        column_config={
            "Güven": st.column_config.ProgressColumn(
                "Güven", min_value=0.0, max_value=100.0, format="%.1f%%"
            ),
            "Model olasılığı": st.column_config.NumberColumn(
                "Yukarı olasılığı", format="%.1f%%"
            ),
            "Test doğruluğu": st.column_config.NumberColumn(
                "Test doğruluğu", format="%.1f%%"
            ),
            "Karşılaştırma puanı": st.column_config.NumberColumn(
                "Beklenen değer", format="%.2f"
            ),
        },
    )
    selected_rows = event.selection.rows
    if selected_rows:
        selected = table.iloc[int(selected_rows[0])]
        st.subheader(f"Seçilen işlem: {selected.get('Varlık', '-')}")
        d1, d2, d3, d4 = st.columns(4)
        d1.metric("Sinyal", str(selected.get("Sinyal", "-")))
        d2.metric("Güven", str(selected.get("Güven yüzdesi", "-")))
        d3.metric(
            "Ödeme",
            f"%{float(selected.get('Binomo ödeme oranı (%)', 0)):.1f}",
        )
        d4.metric(
            "Karşılaştırma",
            f"{float(selected.get('Karşılaştırma puanı', 0)):.2f}",
        )
    return event


@st.fragment(run_every="40s")
def render_auto_top_signal():
    """En az 10 farklı varlığı, 40 saniyede bir otomatik yenileyerek gösterir."""
    if not OUTPUT_FILE.exists():
        st.info("Henüz kaydedilmiş bir AI işlemi bulunmuyor.")
        return
    history = pd.read_excel(OUTPUT_FILE, sheet_name="Sinyaller")
    history = replace_legacy_wait_signals(history)
    if history.empty or "Güven" not in history:
        st.info("Henüz karşılaştırılabilir işlem bulunmuyor.")
        return
    history["Güven"] = pd.to_numeric(history["Güven"], errors="coerce")
    history = history.dropna(subset=["Güven"])
    if history.empty:
        return
    # Her para biriminin yalnızca en güncel sonucunu kullan; aynı varlığın eski
    # kayıtları listenin on farklı öneriye ulaşmasını engellemesin.
    history = history.sort_values("Zaman", ascending=False)
    ranked = (
        history.drop_duplicates(subset=["Varlık"], keep="first")
        .sort_values(["Güven", "Zaman"], ascending=[False, False])
        .reset_index(drop=True)
    )
    page_size = 10
    page_count = max(1, int(np.ceil(len(ranked) / page_size)))
    page_index = (int(time.time()) // 40) % page_count
    start = page_index * page_size
    visible = ranked.iloc[start:start + page_size].copy()
    # Son sayfada 10 satırdan az kaldığında listenin başındaki en yüksek
    # yüzdeli varlıklarla tamamla. Varlık sayısı 10'dan azsa sahte öneri üretme.
    if len(visible) < page_size and len(ranked) >= page_size:
        visible = pd.concat([
            visible,
            ranked.iloc[:page_size - len(visible)],
        ], ignore_index=True)
    visible["Güven yüzdesi"] = visible["Güven"].map(
        lambda value: f"%{float(value) * 100:.1f}"
    )
    visible.insert(0, "Sıra", np.arange(1, len(visible) + 1))
    columns = [
        "Sıra", "Varlık", "Sinyal", "Güven yüzdesi",
        "Binomo ödeme oranı (%)", "Zaman",
    ]
    st.subheader("🏆 Otomatik yüksek güvenli 10 para birimi")
    st.dataframe(
        visible[[column for column in columns if column in visible.columns]],
        use_container_width=True,
        hide_index=True,
    )
    if len(ranked) < page_size:
        st.warning(
            f"Şu anda {len(ranked)} farklı varlık için kayıt var. "
            "10 öneriye ulaşmak için farklı para birimlerinin analiz edilmesi gerekir."
        )
    st.caption(
        f"40 saniyede bir otomatik değişir · Grup {page_index + 1}/{page_count} · "
        "Tıklama gerekmez · En yüksek güven yüzdeleri önceliklidir · "
        + datetime.now().astimezone().strftime("%H:%M:%S")
    )


@st.fragment(run_every="40s")
def render_best_accuracy_panel():
    """Test başarısına göre sıralı modelleri 40 saniyede bir gösterir."""
    if not OUTPUT_FILE.exists():
        return
    history = pd.read_excel(OUTPUT_FILE, sheet_name="Sinyaller")
    accuracy_column = (
        "En başarılı model doğruluğu"
        if "En başarılı model doğruluğu" in history
        else "Test doğruluğu"
    )
    if history.empty or accuracy_column not in history:
        return
    history[accuracy_column] = pd.to_numeric(
        history[accuracy_column], errors="coerce"
    )
    history = history.dropna(subset=[accuracy_column])
    if history.empty:
        return
    ranked = history.sort_values(
        [accuracy_column, "Zaman"], ascending=[False, False]
    ).head(20).reset_index(drop=True)
    rotation_index = (int(time.time()) // 40) % len(ranked)
    best = ranked.iloc[rotation_index]
    accuracy_percent = float(best[accuracy_column]) * 100
    st.subheader("🧠 En çok bilen model")
    b1, b2, b3, b4 = st.columns(4)
    b1.metric("Varlık", str(best.get("Varlık", "-")))
    b2.metric(
        "Model",
        str(best.get("En başarılı model", best.get("Modeller", "-"))),
    )
    b3.metric("Test başarısı", f"%{accuracy_percent:.1f}")
    b4.metric("Sinyal", str(best.get("Sinyal", "-")))
    st.progress(min(max(accuracy_percent / 100, 0.0), 1.0))
    st.caption(
        f"40 saniyede bir değişir · Sıra {rotation_index + 1}/{len(ranked)} · "
        "Yüksek başarı yüzdeleri önceliklidir · "
        + datetime.now().astimezone().strftime("%H:%M:%S")
    )


def latest_indicators(frame: pd.DataFrame) -> dict:
    row = frame.iloc[-1]
    return {
        "EMA 15": round(float(row["close"] / (1 + row["ema_ratio_15"])), 6),
        "EMA15 eğimi": round(float(row["ema_15_slope"]), 6),
        "RSI 14": round(float(row["rsi_14"] * 100), 2),
        "Bollinger %B": round(float(row["bb_percent_b"]), 4),
        "Bollinger genişliği": round(float(row["bb_width"]), 6),
        "MACD": round(float(row["macd"]), 6),
        "MACD histogram": round(float(row["macd_histogram"]), 6),
        "ATR %": round(float(row["atr_14"] * 100), 4),
        "ADX": round(float(row["adx_14"] * 100), 2),
        "Stochastic %K": round(float(row["stoch_k"] * 100), 2),
        "Williams %R": round(float(row["williams_r"] * 100), 2),
        "CCI 20": round(float(row["cci_20"]), 2),
        "MFI 14": round(float(row["mfi_14"] * 100), 2),
    }


def enterprise_context(
    frame: pd.DataFrame,
    probability: float,
    metrics: dict,
) -> dict:
    """Kararı veri kalitesi, rejim ve belirsizlik açısından denetler."""
    feature_frame = frame[FEATURES].replace([np.inf, -np.inf], np.nan)
    missing_ratio = float(feature_frame.isna().mean().mean())
    quality_score = max(0.0, min(100.0, (1 - missing_ratio) * 100))
    latest = frame.iloc[-1]
    adx = float(latest.get("adx_14", 0) * 100)
    volatility = float(latest.get("volatility_20", 0))
    median_volatility = float(
        frame["volatility_20"].dropna().median()
        if frame["volatility_20"].notna().any() else 0
    )
    if adx >= 25:
        regime = "Güçlü trend"
    elif volatility > median_volatility * 1.5:
        regime = "Yüksek volatilite"
    else:
        regime = "Yatay / zayıf trend"
    disagreement = float(metrics.get("Model anlaşmazlığı", 0))
    adaptive_threshold = min(0.70, 0.55 + disagreement)
    directional_confidence = max(probability, 1 - probability)
    # Her taramada mutlaka yön üret. Kalite veya güven düşükse yönü gizlemek
    # yerine açıklamaya düşük güven uyarısı ekle.
    decision = "YUKARI" if probability >= 0.5 else "AŞAĞI"
    if quality_score < 90:
        reason = "Düşük veri kalitesiyle üretilmiş araştırma yönü"
    elif directional_confidence < adaptive_threshold:
        reason = "Dinamik güven eşiğinin altında araştırma yönü"
    else:
        reason = "Kalite ve model uzlaşması yeterli"
    audit_payload = (
        f"{frame['time'].iloc[-1]}|{probability:.8f}|{decision}|"
        f"{quality_score:.4f}|{regime}"
    )
    audit_id = hashlib.sha256(audit_payload.encode("utf-8")).hexdigest()[:16]
    return {
        "decision": decision,
        "reason": reason,
        "quality_score": quality_score,
        "regime": regime,
        "adaptive_threshold": adaptive_threshold,
        "directional_confidence": directional_confidence,
        "audit_id": audit_id,
    }


@st.fragment(run_every="20s")
def render_live_market(asset_name: str, symbol: str, interval: str):
    """Son fiyat ve hızlı teknik göstergeleri sürekli günceller."""
    try:
        live_data = download_market_data(symbol, interval)
        featured = add_features(live_data.tail(250).reset_index(drop=True), 1)
        latest = featured.iloc[-1]
        previous_close = float(live_data["close"].iloc[-2])
        last_close = float(live_data["close"].iloc[-1])
        change = (
            (last_close / previous_close - 1) * 100
            if previous_close else 0.0
        )
        rsi_value = float(latest["rsi_14"] * 100)
        ema15 = float(
            latest["close"] / (1 + latest["ema_ratio_15"])
        )
        technical_score = sum([
            last_close > ema15,
            rsi_value >= 50,
            float(latest["macd_histogram"]) >= 0,
        ])
        technical_direction = (
            "YUKARI" if technical_score >= 2 else "AŞAĞI"
        )
        st.subheader("Canlı piyasa verisi")
        l1, l2, l3, l4, l5 = st.columns(5)
        l1.metric("Varlık", asset_name)
        l2.metric("Son fiyat", f"{last_close:.6f}", f"{change:+.3f}%")
        l3.metric("RSI 14", f"{rsi_value:.1f}")
        l4.metric("EMA 15", f"{ema15:.6f}")
        l5.metric("Hızlı teknik yön", technical_direction)
        st.caption(
            f"{interval} veri · 20 saniyede bir kontrol · "
            + datetime.now().astimezone().strftime("%H:%M:%S")
        )
    except Exception as exc:
        st.warning(f"Canlı veri geçici olarak yenilenemedi: {exc}")


@st.fragment(run_every="40s")
def auto_advance_market():
    """Her 40 saniyede sıradaki varlığa geçip tam analizi yeniden çalıştırır."""
    now = time.time()
    last_change = st.session_state.get("auto_market_last_change")
    if last_change is None:
        st.session_state.auto_market_last_change = now
        return
    if now - float(last_change) >= 38:
        st.session_state.auto_market_index = (
            int(st.session_state.get("auto_market_index", 0)) + 1
        ) % len(AUTO_SCAN_ASSETS)
        st.session_state.auto_market_last_change = now
        st.rerun()


st.set_page_config(page_title="Binomo DL Araştırma", layout="wide")
st.title("Binomo Derin Öğrenme Araştırması")
st.caption("Sürüm: ENTERPRISE-AI-2026.07.30.16 — Governed Intelligence")
st.warning("Araştırma amaçlıdır. Otomatik işlem yapmaz; yatırım tavsiyesi veya kazanç garantisi değildir.")
render_auto_top_signal()
render_best_accuracy_panel()
auto_advance_market()
data_source = "Çevrim içi veriyi kendisi indir"
uploaded = None
interval = "15m"
st.info(
    "Tam otomatik tarama etkin: Her 40 saniyede sıradaki varlık analiz edilir "
    "ve sonuç Excel'e kaydedilir. Binomo OTC seçimlerinde normal piyasa "
    "karşılığı (vekil sembol) kullanılır."
)
c1, c2, c3 = st.columns(3)
if data_source == "Çevrim içi veriyi kendisi indir":
    market_names = AUTO_SCAN_ASSETS
    market_index = int(st.session_state.get("auto_market_index", 0)) % len(market_names)
    asset = market_names[market_index]
    c1.metric("Otomatik taranan varlık", asset)
    interval = c2.selectbox("Mum aralığı", list(INTERVAL_PERIODS), index=2)
    horizon = c3.number_input("Tahmin ufku (mum)", 1, 20, 1)
else:
    asset = c1.text_input("Varlık", "EUR/USD")
    horizon = c2.number_input("Tahmin ufku (mum)", 1, 20, 1)
    interval = c3.text_input("Mum aralığı", "CSV")
if data_source == "Çevrim içi veriyi kendisi indir":
    render_live_market(asset, MARKET_SYMBOLS[asset], interval)
c4, c5, c6 = st.columns(3)
lookback = c4.number_input("Model penceresi (mum)", 20, 120, 40)
epochs = c5.slider("DL epoch", 5, 60, 15)
payout_rate = c6.number_input(
    "Binomo ödeme oranı (%)", min_value=0, max_value=100, value=82,
    help="Platformda görünen oranı elle girin; model güveninden farklıdır.",
)
analysis_candles = st.slider(
    "Analiz edilecek son mum sayısı",
    min_value=1000,
    max_value=30000,
    value=30000,
    step=1000,
    help=(
        "Göstergeler tüm mevcut mumlarda hesaplanır. Model eğitimi Cloud "
        "belleği için zaman geneline yayılmış en fazla 6.000 pencere kullanır."
    ),
)
csv_bytes = None
csv_source = None
online_data = None
if data_source == "Çevrim içi veriyi kendisi indir":
    try:
        online_data = download_market_data(MARKET_SYMBOLS[asset], interval)
        csv_bytes = online_data.to_csv(index=False).encode("utf-8")
        csv_source = f"{MARKET_SYMBOLS[asset]} ({interval})"
        st.caption(f"{len(online_data):,} mum otomatik indirildi.")
    except Exception as exc:
        st.error(f"Çevrim içi veri alınamadı: {exc}")
elif uploaded is not None:
    csv_bytes = uploaded.getvalue()
    csv_source = uploaded.name
elif AUTO_CSV_FILE.exists():
    csv_bytes = AUTO_CSV_FILE.read_bytes()
    csv_source = AUTO_CSV_FILE.name
    st.info(f"Hazır veri bulundu: {AUTO_CSV_FILE.name}")

analysis_key = None
if csv_bytes:
    settings = (
        f"{asset}|{horizon}|{lookback}|{epochs}|{analysis_candles}"
    ).encode("utf-8")
    analysis_key = hashlib.sha256(csv_bytes + settings).hexdigest()

if "last_saved_analysis" not in st.session_state:
    st.session_state.last_saved_analysis = None

# Veri hazır olur olmaz analiz otomatik başlar. Aynı veri ve ayarlar, oturum
# içinde yeniden eğitilmez; böylece 40 saniyelik ekran yenilemeleri ağır modeli
# gereksiz yere tekrar çalıştırmaz.
should_analyze = bool(
    csv_bytes and st.session_state.last_saved_analysis != analysis_key
)

if should_analyze:
    try:
        if online_data is not None:
            data = online_data
        else:
            raw = pd.read_csv(io.BytesIO(csv_bytes), sep=None, engine="python")
            data = normalize_columns(raw)
        selected_data = data.tail(int(analysis_candles)).reset_index(drop=True)
        if len(selected_data) < int(analysis_candles):
            st.warning(
                f"Veri kaynağı {analysis_candles:,} yerine yalnızca "
                f"{len(selected_data):,} gerçek mum sağladı. Analiz mevcut "
                "mumlarla yapılıyor; yapay mum eklenmiyor."
            )
        frame = add_features(selected_data, int(horizon))
        with st.spinner("Modeller eğitiliyor..."):
            probability, metrics, model_names, model_details = train_and_predict(
                frame, int(lookback), epochs, int(analysis_candles)
            )
        indicators = latest_indicators(frame)
        enterprise = enterprise_context(frame, probability, metrics)
        signal = enterprise["decision"]
        confidence = enterprise["directional_confidence"]
        best_model_name = max(
            model_details,
            key=lambda name: model_details[name]["test_doğruluğu"],
        )
        best_model_accuracy = model_details[best_model_name]["test_doğruluğu"]
        record = {
            "Zaman": datetime.now().astimezone().isoformat(timespec="seconds"),
            "Varlık": asset,
            "Sinyal": signal,
            "Model olasılığı": round(probability, 4),
            "Güven": round(confidence, 4),
            "Güven yüzdesi": f"%{confidence * 100:.1f}",
            "Binomo ödeme oranı (%)": payout_rate,
            "Tahmin ufku (mum)": horizon,
            "Model sayısı": len(model_names),
            "Modeller": ", ".join(model_names),
            "En başarılı model": best_model_name,
            "En başarılı model doğruluğu": best_model_accuracy,
            "AI karar nedeni": (
                enterprise["reason"]
            ),
            "Piyasa rejimi": enterprise["regime"],
            "Veri kalite puanı": round(enterprise["quality_score"], 2),
            "Dinamik güven eşiği": round(
                enterprise["adaptive_threshold"], 4
            ),
            "Denetim kimliği": enterprise["audit_id"],
            "Test doğruluğu": round(metrics["Doğruluk"], 4),
            "Test precision": round(metrics["Precision"], 4),
            "Test recall": round(metrics["Recall"], 4),
            "İndirilen mum sayısı": len(data),
            "Analiz edilen mum sayısı": len(selected_data),
            **indicators,
        }
        excel_bytes = append_excel(record, selected_data)
        st.session_state.last_saved_analysis = analysis_key

        # Kullanıcının ilk gördüğü bölüm: önceliklendirilmiş işlem listesi.
        st.subheader("Öncelikli işlemler")
        st.caption("En yüksek güven yüzdesi ilk sıradadır.")
        priority_table = prioritized_signal_table()
        render_clickable_ranking(priority_table, "priority_after_analysis")
        st.success(
            f"{csv_source} otomatik analiz edildi ve Excel'e kaydedildi. "
            f"Sonuç: {signal} — model güveni %{confidence * 100:.1f}"
        )
        st.json(metrics)
        st.subheader("Enterprise karar denetimi")
        e1, e2, e3, e4 = st.columns(4)
        e1.metric("Piyasa rejimi", enterprise["regime"])
        e2.metric(
            "Veri kalitesi", f"%{enterprise['quality_score']:.1f}"
        )
        e3.metric(
            "Dinamik eşik",
            f"%{enterprise['adaptive_threshold'] * 100:.1f}",
        )
        e4.metric("Denetim ID", enterprise["audit_id"])
        st.info("Karar açıklaması: " + enterprise["reason"])
        with st.expander("Zeka motoru — model ağırlıkları"):
            st.dataframe(
                pd.DataFrame(model_details).T.reset_index(
                    names="Model"
                ),
                use_container_width=True,
                hide_index=True,
            )
        st.subheader("Son mum teknik göstergeleri")
        st.dataframe(pd.DataFrame([indicators]), use_container_width=True, hide_index=True)
        st.caption("Kullanılan modeller: " + ", ".join(model_names))
        st.download_button(
            "Excel'i indir", excel_bytes, OUTPUT_FILE.name,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as exc:
        logging.exception("Analiz başarısız")
        st.error(f"İşlem tamamlanamadı: {exc}")
elif csv_bytes and st.session_state.last_saved_analysis == analysis_key:
    st.success("Bu veri daha önce analiz edilip Excel'e kaydedildi.")
    existing = prioritized_signal_table()
    if not existing.empty:
        st.subheader("Öncelikli işlemler")
        render_clickable_ranking(existing, "priority_existing")
elif csv_bytes:
    st.info("Piyasa verisi hazır; analiz otomatik olarak başlatılıyor.")
